import re

import pandas as pd
import openpyxl
import numpy as np


class IOData:

    @staticmethod
    def get_data():
        xls = pd.ExcelFile('data/input_data.xlsx')
        all_input = []
        for name in xls.sheet_names:
            all_input.append(pd.read_excel('data/input_data.xlsx', index_col=None, header=None, sheet_name=name).to_numpy())
        return all_input

    @staticmethod
    def output_data(data1, data2, n):
        wb = openpyxl.load_workbook('data/output_data.xlsx')
        new_sheet = re.sub(r'[0-9]+$',
                           lambda x: f"{str(int(x.group()) + 1).zfill(len(x.group()))}",
                           wb.sheetnames[-1])
        sheet = wb.create_sheet(new_sheet)

        sheet.merge_cells('B1:D1')
        sheet.merge_cells('A1:A2')
        sheet.cell(row=1, column=1).value = "Кількість елементів"
        sheet.cell(row=1, column=2).value = "Середній час"
        sheet.cell(row=2, column=2).value = "B&B"
        sheet.cell(row=2, column=3).value = "Greed"
        sheet.cell(row=2, column=4).value = "ACO"

        sheet.column_dimensions['A'].width = 20
        for i in range(1, len(n) + 1):
            sheet.cell(row=i + 2, column=1).value = n[i - 1]

        for i in range(len(n)):
            for j in range(3):
                sheet.cell(row=i + 3, column=j + 2).value = data1[i][j]

        sheet.merge_cells('G1:I1')
        sheet.merge_cells('F1:F2')
        sheet.cell(row=1, column=6).value = "Кількість елементів"
        sheet.cell(row=1, column=7).value = "Точність"
        sheet.cell(row=2, column=7).value = "B&B"
        sheet.cell(row=2, column=8).value = "Greed"
        sheet.cell(row=2, column=9).value = "ACO"

        sheet.column_dimensions['F'].width = 20
        for i in range(1, len(n) + 1):
            sheet.cell(row=i + 2, column=6).value = n[i - 1]
        for i in range(len(n)):
            for j in range(3):
                sheet.cell(row=i + 3, column=j + 7).value = data2[i][j]

        wb.save('data/output_data.xlsx')


if __name__ == "__main__":
    a = IOData.get_data()
    print(a[1])
    print(
       IOData.output_data([[3, 3, 4], [4, 5, 6], [1, 2, 3], [5, 6, 4]],
                          [[1, 0.7, 0.8], [1, 0.78, 0.89], [1, 0.75, 0.85], [1, 0.73, 0.887]],
                           [5, 10, 20, 30]))
